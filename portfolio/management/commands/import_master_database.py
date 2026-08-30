import os
import re
import openpyxl
from decimal import Decimal
from django.core.management.base import BaseCommand
from django.utils.text import slugify
from django.conf import settings

from portfolio.models import Brand, Vehicle
from calculator.models import State, VehicleStateEstimate

# State Name to 2-letter Code Mapping for Indian States in Master Database
STATE_CODE_MAP = {
    "Andhra Pradesh": "AP",
    "Arunachal Pradesh": "AR",
    "Assam": "AS",
    "Bihar": "BR",
    "Chhattisgarh": "CG",
    "Goa": "GA",
    "Gujarat": "GJ",
    "Haryana": "HR",
    "Himachal Pradesh": "HP",
    "Jharkhand": "JH",
    "Karnataka": "KA",
    "Kerala": "KL",
    "Madhya Pradesh": "MP",
    "Maharashtra": "MH",
    "Manipur": "MN",
    "Meghalaya": "ML",
    "Mizoram": "MZ",
    "Nagaland": "NL",
    "Odisha": "OD",
    "Punjab": "PB",
    "Rajasthan": "RJ",
    "Sikkim": "SK",
    "Tamil Nadu": "TN",
    "Telangana": "TG",
    "Tripura": "TR",
    "Uttar Pradesh": "UP",
    "Uttarakhand": "UK",
    "West Bengal": "WB",
}

def parse_currency(val):
    if val is None or str(val).strip() == "" or str(val).strip() == "TBA":
        return None
    val_str = str(val).replace('₹', '').replace(',', '').strip()
    if val_str == "TBA" or val_str == "":
        return None
    
    # Matches e.g. 4.25 Cr, 46.99 L, 85 L, 50 K
    m_cr = re.search(r'([\d.]+)\s*Cr', val_str, re.IGNORECASE)
    if m_cr:
        return Decimal(int(round(float(m_cr.group(1)) * 10_000_000)))
    
    m_l = re.search(r'([\d.]+)\s*L', val_str, re.IGNORECASE)
    if m_l:
        return Decimal(int(round(float(m_l.group(1)) * 100_000)))
        
    m_k = re.search(r'([\d.]+)\s*K', val_str, re.IGNORECASE)
    if m_k:
        return Decimal(int(round(float(m_k.group(1)) * 1000)))

    try:
        return Decimal(int(round(float(val_str))))
    except ValueError:
        return None


class Command(BaseCommand):
    help = 'Import 2026 India Car Master Database from XLSX files into Django models'

    def add_arguments(self, parser):
        parser.add_argument(
            '--catalog',
            type=str,
            default='India_Car_Master_Database_August_2026.xlsx',
            help='Path to master catalog Excel file'
        )
        parser.add_argument(
            '--statewise',
            type=str,
            default='India_Car_Master_Database_2026_Statewise_OnRoad_Prices.xlsx',
            help='Path to statewise prices Excel file'
        )
        parser.add_argument(
            '--force',
            action='store_true',
            help='Overwrite manual edits'
        )

    def handle(self, *args, **options):
        catalog_path = options['catalog']
        statewise_path = options['statewise']
        force = options['force']

        # Find absolute paths if relative
        workspace_dir = settings.BASE_DIR.parent
        if not os.path.isabs(catalog_path):
            if os.path.exists(os.path.join(workspace_dir, catalog_path)):
                catalog_path = os.path.join(workspace_dir, catalog_path)
            elif os.path.exists(os.path.join(settings.BASE_DIR, catalog_path)):
                catalog_path = os.path.join(settings.BASE_DIR, catalog_path)
            else:
                catalog_path = os.path.abspath(catalog_path)

        if not os.path.isabs(statewise_path):
            if os.path.exists(os.path.join(workspace_dir, statewise_path)):
                statewise_path = os.path.join(workspace_dir, statewise_path)
            elif os.path.exists(os.path.join(settings.BASE_DIR, statewise_path)):
                statewise_path = os.path.join(settings.BASE_DIR, statewise_path)
            else:
                statewise_path = os.path.abspath(statewise_path)

        self.stdout.write(self.style.SUCCESS(f"Starting import from:\n  Catalog: {catalog_path}\n  Statewise: {statewise_path}"))

        if not os.path.exists(catalog_path):
            self.stderr.write(self.style.ERROR(f"Catalog file not found at {catalog_path}"))
            return
        if not os.path.exists(statewise_path):
            self.stderr.write(self.style.ERROR(f"Statewise file not found at {statewise_path}"))
            return

        # ----------------------------------------------------
        # 1. PARSE CATALOG FILE
        # ----------------------------------------------------
        wb_cat = openpyxl.load_workbook(catalog_path, data_only=True)
        ws_cat = wb_cat['Master Database'] if 'Master Database' in wb_cat.sheetnames else wb_cat.active

        catalog_rows = list(ws_cat.iter_rows(values_only=True))[1:]
        self.stdout.write(f"Loaded catalog sheet: {len(catalog_rows)} vehicle rows found.")

        brand_count = 0
        vehicle_count = 0
        flagged_vehicles = []
        vehicle_map = {} # (brand_name, car_name) -> Vehicle instance

        for row_idx, r in enumerate(catalog_rows, start=2):
            if not r or not r[0]:
                continue
            
            brand_name = str(r[0]).strip()
            car_name = str(r[1]).strip()
            body_type = str(r[2]).strip() if r[2] else 'SUV'
            fuel_type = str(r[3]).strip() if r[3] else 'Petrol'
            ev_flag = str(r[4]).strip() if r[4] else 'No'
            start_price_raw = r[5]
            top_price_raw = r[6]
            seats_raw = r[7]
            transmission_raw = str(r[8]).strip() if r[8] else 'Manual'

            # Currency parsing
            start_price = parse_currency(start_price_raw)
            top_price = parse_currency(top_price_raw)

            needs_review = False
            review_reasons = []

            # Check TBA prices
            if start_price is None or top_price is None or start_price_raw == "TBA" or top_price_raw == "TBA":
                needs_review = True
                review_reasons.append("TBA / Missing price")

            # Check Seats validation
            seats = None
            if seats_raw is not None:
                try:
                    val = float(seats_raw)
                    if val.is_integer() and val > 0:
                        seats = int(val)
                    else:
                        needs_review = True
                        review_reasons.append(f"Invalid decimal seats count ({seats_raw})")
                except (ValueError, TypeError):
                    needs_review = True
                    review_reasons.append(f"Unparseable seats count ({seats_raw})")

            # Check Transmission TBA
            if transmission_raw == "TBA":
                needs_review = True
                review_reasons.append("Transmission TBA")

            # Upsert Brand
            brand_slug = slugify(brand_name)
            brand, b_created = Brand.objects.get_or_create(
                name=brand_name,
                defaults={'slug': brand_slug}
            )
            if b_created:
                brand_count += 1

            # Determine vehicle active status
            is_active = not needs_review

            # Upsert Vehicle
            vehicle, v_created = Vehicle.objects.get_or_create(
                brand=brand,
                name=car_name,
                defaults={
                    'body_type': body_type,
                    'fuel_type': fuel_type,
                    'ev_hybrid_cng_flag': ev_flag,
                    'starting_price': start_price or Decimal('0'),
                    'top_variant_price': top_price or Decimal('0'),
                    'ex_showroom_price': start_price or Decimal('0'),
                    'seats': seats,
                    'transmission': transmission_raw,
                    'needs_review': needs_review,
                    'is_active': is_active,
                    'data_source': 'master_db_import',
                    'meta_title': f"{brand_name} {car_name} Price, Specs & On-Road Calculator | Car Guide Media",
                    'meta_description': f"Calculate exact on-road price for {brand_name} {car_name} across all Indian states and union territories."
                }
            )

            # Update vehicle if not manual edit or force flag set
            if not v_created:
                if vehicle.data_source == 'master_db_import' or force:
                    vehicle.body_type = body_type
                    vehicle.fuel_type = fuel_type
                    vehicle.ev_hybrid_cng_flag = ev_flag
                    vehicle.starting_price = start_price or Decimal('0')
                    vehicle.top_variant_price = top_price or Decimal('0')
                    vehicle.ex_showroom_price = start_price or Decimal('0')
                    vehicle.seats = seats
                    vehicle.transmission = transmission_raw
                    vehicle.needs_review = needs_review
                    vehicle.is_active = is_active
                    vehicle.data_source = 'master_db_import'
                    vehicle.save()

            if v_created:
                vehicle_count += 1

            if needs_review:
                flagged_vehicles.append((row_idx, brand_name, car_name, ", ".join(review_reasons)))

            vehicle_map[(brand_name, car_name)] = vehicle

        # ----------------------------------------------------
        # 2. PARSE STATE NOTES SHEET
        # ----------------------------------------------------
        wb_state = openpyxl.load_workbook(statewise_path, data_only=True)
        
        state_notes = {}
        if 'State Notes' in wb_state.sheetnames:
            ws_notes = wb_state['State Notes']
            for row in list(ws_notes.iter_rows(values_only=True))[1:]:
                if row and row[0]:
                    st_name = str(row[0]).strip()
                    note_text = str(row[1]).strip() if len(row) > 1 and row[1] else "Estimated 2026 road-tax basis; see PDF methodology."
                    state_notes[st_name] = note_text

        # ----------------------------------------------------
        # 3. PARSE STATEWISE PRICES SHEET
        # ----------------------------------------------------
        ws_prices = wb_state['Statewise Prices'] if 'Statewise Prices' in wb_state.sheetnames else wb_state.active
        headers = [cell.value for cell in ws_prices[1]]

        # Map state headers to column pairs
        state_cols = {} # state_name -> {'start': col_idx, 'top': col_idx}
        for col_idx, h in enumerate(headers):
            if h and ("Start OTR" in str(h) or "Top OTR" in str(h)):
                parts = str(h).split()
                # e.g. "Andhra Pradesh Start OTR" -> "Andhra Pradesh"
                # e.g. "West Bengal Top OTR" -> "West Bengal"
                if "Start OTR" in str(h):
                    st_name = str(h).replace("Start OTR", "").strip()
                    if st_name not in state_cols:
                        state_cols[st_name] = {}
                    state_cols[st_name]['start'] = col_idx
                elif "Top OTR" in str(h):
                    st_name = str(h).replace("Top OTR", "").strip()
                    if st_name not in state_cols:
                        state_cols[st_name] = {}
                    state_cols[st_name]['top'] = col_idx

        # Ensure State objects exist
        states_processed = 0
        state_obj_map = {}
        for st_name, cols in state_cols.items():
            code = STATE_CODE_MAP.get(st_name, st_name[:2].upper())
            note = state_notes.get(st_name, "Estimated 2026 road-tax basis; see PDF methodology.")
            state_obj, s_created = State.objects.get_or_create(
                name=st_name,
                defaults={
                    'code': code,
                    'data_source_note': note,
                    'is_active': True
                }
            )
            if not s_created and state_obj.data_source_note != note:
                state_obj.data_source_note = note
                state_obj.save()

            state_obj_map[st_name] = state_obj
            states_processed += 1

        # Process Vehicle State Estimates
        estimate_count = 0
        price_rows = list(ws_prices.iter_rows(values_only=True))[1:]

        for row in price_rows:
            if not row or not row[0] or not row[1]:
                continue
            brand_name = str(row[0]).strip()
            car_name = str(row[1]).strip()

            vehicle = vehicle_map.get((brand_name, car_name))
            if not vehicle:
                continue

            for st_name, cols in state_cols.items():
                state_obj = state_obj_map.get(st_name)
                if not state_obj:
                    continue

                start_col = cols.get('start')
                top_col = cols.get('top')

                val_start_raw = row[start_col] if start_col is not None and start_col < len(row) else None
                val_top_raw = row[top_col] if top_col is not None and top_col < len(row) else None

                parsed_start_otr = parse_currency(val_start_raw)
                parsed_top_otr = parse_currency(val_top_raw)

                if parsed_start_otr is not None and parsed_top_otr is not None:
                    VehicleStateEstimate.objects.update_or_create(
                        vehicle=vehicle,
                        state=state_obj,
                        defaults={
                            'start_otr': parsed_start_otr,
                            'top_otr': parsed_top_otr
                        }
                    )
                    estimate_count += 1

        # ----------------------------------------------------
        # 4. PRINT SUMMARY REPORT
        # ----------------------------------------------------
        self.stdout.write(self.style.SUCCESS("\n=================================================="))
        self.stdout.write(self.style.SUCCESS(" MASTER DATABASE IMPORT SUMMARY REPORT "))
        self.stdout.write(self.style.SUCCESS("=================================================="))
        self.stdout.write(f"Brands Created/Verified: {Brand.objects.count()}")
        self.stdout.write(f"Vehicles Processed: {Vehicle.objects.count()} (New: {vehicle_count})")
        self.stdout.write(f"Active Vehicles: {Vehicle.objects.filter(is_active=True).count()}")
        self.stdout.write(self.style.WARNING(f"Flagged for Review (needs_review=True): {len(flagged_vehicles)}"))
        self.stdout.write(f"States Processed: {states_processed}")
        self.stdout.write(f"Vehicle State Estimates Created/Updated: {estimate_count}")
        self.stdout.write("--------------------------------------------------")
        self.stdout.write(self.style.WARNING("FLAGGED VEHICLES LIST:"))
        for f in flagged_vehicles:
            self.stdout.write(f"  Row {f[0]}: {f[1]} {f[2]} -> Reasons: {f[3]}")
        self.stdout.write(self.style.SUCCESS("==================================================\n"))

